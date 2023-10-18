import pandas as pd 
import openpyxl
import os,re,datetime
from openpyxl.utils import get_column_letter


d = {
	"BV " : ['V2', 2, 1] ,
	"VGSTH " : ['Vgsth', 3, 1],
	"IGSS " : ['Ig', 4, 1E9],
	"Idss 1200V " : ['IDSS', 5, 1E9],
	"Rdson 18V " : ['Rdson', 6, 1E3]
}
path = r'D:\可靠性实验\项目\体二极管双极退化实验\BD\0MIN'
row_dict = {}

wb = openpyxl.Workbook()
for file in os.listdir(path):

	if file[-3:] not in ['csv', 'CSV', 'xls']:
		continue

	# BV [TO3P-3VD-03A(21); 12_8_2022 5_29_31 PM]
	col_step = 0
	test = re.match(r".*(?=\[)", file, re.I).group()
	index = int(re.findall(r'[(](.*?)[)]', file)[0])
	project_name = re.findall(r"\[(.*)\(", file, re.I)[0]

	if project_name in wb.sheetnames:
		ws = wb[project_name]
	else:
		ws = wb.create_sheet(project_name, len(wb.sheetnames))
		# ws.cell(1, 1).value = 'INDEX'
		title_list = ['INDEX'] + list(d.keys()) + ['BD 0V', 'BD -2V', 'BD -4V']
		for col in range(len(title_list)):
			ws.cell(1, col+1).value = title_list[col]
		# for name, col in zip(d.keys(), range(1, len(d)+1)):
		# 	ws.cell(1, col).value = name

	df = pd.read_csv(os.path.join(path, file))
	ws.cell(index+1, 1).value = index

	if test == 'Body charasteristics ':
		ws.cell(index+1, 7).value = df.loc[(df['Id'] == -20)&(df['Vg'] == 0), 'Vd'].tolist()[0]
		ws.cell(index+1, 8).value = df.loc[(df['Id'] == -20)&(df['Vg'] == -2), 'Vd'].tolist()[0]
		ws.cell(index+1, 9).value = df.loc[(df['Id'] == -20)&(df['Vg'] == -4), 'Vd'].tolist()[0]
	else:
		try:
			ws.cell(index+1, d[test][1]).value = float(df[d[test][0]][0]) * d[test][2]
		except:
			ws.cell(index+1, d[test][1]).value = '-'
save_path = r'C:\Users\Wufanzhengshu\Desktop'
date = datetime.datetime.today().strftime('%Y%m%d_%H%M%S')
wb.save(os.path.join(save_path, f'summary_{project_name}_{date}.xlsx'))
# if os.path.isfile(project_name):
# 	wb.save(project_name + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + '.xlsx')
# else:
# 	wb.save(r'C:\Users\Wufanzhengshu\Desktop\rest.xlsx')
		
