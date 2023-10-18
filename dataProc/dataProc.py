import re, csv, configparser, json, re, xlrd, xlwt, xlutils, os
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook

class BaseObject():

	def __init__(self):
		pass

	def getColName(self, file, header):	

		data_iter = csv.reader(open(file, 'r', encoding = 'gbk'))
		data = next(data_iter)
		index_dict = {	'No.' : [], 
						'Conditon1' : [], 
						'Conditon2' : [], 
						'Conditon3' : []}

		for i in range(0, header):
			if len(data) > 0 and data[0] in index_dict.keys():
				index_dict[data[0]] = data
			data = next(data_iter)

		index_dict['No.'][0] = 'INDEX'

		for i in range(4, len(index_dict['No.'])):

			if len(index_dict['Conditon1'][i]) > 0 and index_dict['Conditon1'][i] != '-':
				index_dict['No.'][i] = index_dict['No.'][i] + '@' + index_dict['Conditon1'][i]
			else:
				continue

			if index_dict['Conditon2'][i] != '-':
				index_dict['No.'][i] = index_dict['No.'][i] + '&' + index_dict['Conditon2'][i]

			if len(index_dict['Conditon3']) > 0 and index_dict['Conditon3'][i] != '-':
				index_dict['No.'][i] = index_dict['No.'][i] + '&' + index_dict['Conditon3'][i]	

		return index_dict['No.']

	def searchHeader(self, file):

		data_iter = csv.reader(open(file, 'r', encoding = 'gbk'))
		data = next(data_iter)
		header_len = 0

		while True:

			try:
				int(data[0])
				data = next(data_iter)
				return header_len
			except (ValueError, IndexError):
				data = next(data_iter)
				header_len += 1
			except StopIteration:
				return header_len

		return header_len

class DataObject(BaseObject):

	def __init__(self, 
				cols_name = ['BVDSS@VDSmax=1800V&IDS=200uA', 
							'VGS(TH)@IDS=10mA',
							'IDSS@VDS=1200V', 
							'IGSS@VGS=18V', 
							'RDS(ON)@IDS=40A&VG=15V']):
		
		self.cols_name = cols_name
		self.wb = Workbook()

		# pd.set_option('display.max_columns', None)
		# pd.set_option('display.max_rows', None)
		
	def save(self, save_path, save_name):
		del self.wb['Sheet']
		# self.wb.save(r"C:/Users/Wufanzhengshu/Desktop/result.xlsx")
		file_path = os.path.join(save_path, save_name)
		self.wb.save(file_path)
		with pd.ExcelWriter(file_path, mode = 'a') as writer:
			self.df.to_excel(writer, sheet_name='DATA')
	
	def sortedCol(self):

		col_list = []
		df = self.df
		second_cols = ['PRE', 'POST', 'RATE']
		exist_cols = df.columns
		for colname in self.cols_name:
			for scol in second_cols:
				if (colname, scol) in exist_cols:
					col_list.append((colname, scol))

		self.df = df[col_list]			
		return df[col_list]

	def addData(self, df_file):
		self.df = pd.read_excel(df_file, sheet_name = 'DATA', index_col=[0,1,2], header = [0,1])
		fileinfo_list = self.filePack(filedialog.askopenfilenames())
		self.loadData(fileinfo_list)
		self.addRate()
		self.sortedCol()

	def loadFromFiles(self, path):
		fileinfo_list = self.filePack(path)
		self.df = pd.DataFrame()
		print(fileinfo_list)
		self.loadData(fileinfo_list)
		self.addRate()
		self.sortedCol()

	def addRate(self):
		cols_name = []
		df = self.df
		wb = self.wb
		
		for colname in self.cols_name:
			if (colname, 'PRE') not in df.columns or (colname, 'POST') not in df.columns:
				continue
			cols_name.append(colname)

		sheet_names = []
		rows_dict = {}
		for index in self.df.index:
			sheet_name = index[0] + '-' + index[1]
			row = index[2]

			if sheet_name not in sheet_names:
				# ws = wb.add_sheet(sheet_name)
				ws = wb.create_sheet(sheet_name)
				# ws.write(0, 0, 'INDEX')
				ws.cell(row=1, column=1, value='INDEX')
				for col, colname in enumerate(self.cols_name):
					# ws.write(0, col+1, colname)
					ws.cell(row=1, column=col+2, value=colname)
				ws.cell(row=2, column=1, value=row)
				sheet_names.append(sheet_name)
				rows_dict[sheet_name] = {}
				rows_dict[sheet_name][row] = 2
			else:
				rows_dict[sheet_name][row] = len(rows_dict[sheet_name])+2
				ws.cell(row=rows_dict[sheet_name][row], column=1, value=row)
				
		for index, df_row in df.iterrows():

			project_name,test_ID,row = index
			sheet_name = project_name + '-' + test_ID
			for col, colname in enumerate(cols_name):

				pre = str(df_row[(colname, 'PRE')])
				post = str(df_row[(colname, 'POST')])
				short_colname = colname.split('@')[0]
				mode = 'STYLE_MULTI' if any(key in short_colname for key in ['IDSS', 'IGSS', 'IR']) else 'STYLE_PERCENT'
				mode_raw = 'RAW_MULTI' if any(key in short_colname for key in ['IDSS', 'IGSS', 'IR']) else 'RAW_PERCENT'
				
				if pd.isnull(pre) or pd.isnull(post):
					df.loc[index, (colname, 'RATE')] = '-'
					ws.cell(row=rows_dict[sheet_name][row], column=col+2, value='-')
				else:
					ws = wb[sheet_name]
					ws.cell(row=rows_dict[sheet_name][row], column=col+2, value=self.calRateByMode(pre, post, mode_raw))
					df.loc[index, (colname, 'RATE')] = self.calRateByMode(pre, post, mode)
		
		return df

	def delFailData(self):

		df = self.df
		project_list = list(df.index.get_level_values('PROJECT').unique())
		test_list = list(df.index.get_level_values('TEST').unique())
		group_flag = 80

		for project in project_list:
			for test in test_list:
				df_test = df.loc[project, test]

				for index, row in df_test:
					pass

	def addOriginData(self):

		cols_name = []
		df = self.df 
		wb = self.wb 
		for colname in self.cols_name:
			if (colname, 'PRE') not in df.columns or (colname, 'POST') not in df.columns:
				continue
			cols_name.append(colname)

	#文件分组函数
	#根据文件名将文件划分到不同组中，后续数据分类对比依据此函数输出的info_dict
	def filePack(self, para):

		if type(para) == list or type(para) == tuple:
			files = para
		elif type(para) == str:
			files = os.listdir(para)
		else:
			raise Exception("File Parameter is WRONG. Please Check It")

		file_list = []
		for file in files:
			if file[-3:] in ['csv', 'CSV']:
				# print("Skip : " + file[-3:])
				file_list.append(file)			

		#todo 文件名合法性验证
		fileinfo_list = []
		for file_name in file_list:

			if '/' in file_name or '\\' in file_name:
				path = os.path.dirname(file_name)
				file_name = os.path.basename(file_name)
			else:
				path = para
				
			#读取一个file_list中的元素，依次读取其中字符，按空格划分关键词
			info_list = []
			info_item = ''
			for i in file_name[:-4]:
				if i == ' ':
					info_list.append(info_item)
					info_item = ''
					continue
				info_item += i
			else:
				info_list.append(info_item)

			#文件名中的关键词依次对应：项目名、试验项、时间节点
			project_name, test_ID, time = info_list

			fileinfo_list.append(
				{
					'PROJECT' : project_name,
					'TEST' : test_ID,
					'TIME' : time,
					'PATH' : path + '/' + file_name
				})

		return fileinfo_list

	def getColName(self, file, header):

		data_iter = csv.reader(open(file, 'r', encoding = 'gbk'))
		data = next(data_iter)
		index_dict = {	'No.' : [], 
						'Conditon1' : [], 
						'Conditon2' : [], 
						'Conditon3' : []}

		for i in range(0, header):
			if len(data) > 0 and data[0] in index_dict.keys():
				index_dict[data[0]] = data
			data = next(data_iter)

		index_dict['No.'][0] = 'INDEX'

		for i in range(4, len(index_dict['No.'])):

			if len(index_dict['Conditon1'][i]) > 0 and index_dict['Conditon1'][i] != '-':
				index_dict['No.'][i] = index_dict['No.'][i] + '@' + index_dict['Conditon1'][i]
			else:
				continue

			if index_dict['Conditon2'][i] != '-':
				index_dict['No.'][i] = index_dict['No.'][i] + '&' + index_dict['Conditon2'][i]

			if len(index_dict['Conditon3']) > 0 and index_dict['Conditon3'][i] != '-':
				index_dict['No.'][i] = index_dict['No.'][i] + '&' + index_dict['Conditon3'][i]	

		return index_dict['No.']

	def getHeader(self, file):

		data_iter = csv.reader(open(file, 'r', encoding = 'gbk'))
		data = next(data_iter)
		header_len = 0
		flag = 2

		while flag > 0:
			if len(data) > 0 and '---' in data[0]:
				flag -= 1
			data = next(data_iter)
			header_len += 1

		while True:
			if len(data) != 0:
				try:
					int(data[0])

					break
				except ValueError:
					pass
			data = next(data_iter)
			header_len += 1

		return header_len

	def loadData(self, fileinfo_list):

		df_dict = {}

		for file in fileinfo_list:
			project_name, test_ID, time, file_path= file.values()

			header_len = self.getHeader(file_path)
			cols_name = self.getColName(file_path, header_len)

			for i in range(4, len(cols_name)):
				if cols_name[i] in cols_name[i+1:]:
					cols_name[i] += str(i)

			df_csv = pd.read_csv(file_path, header = header_len - 1, names = cols_name, skip_blank_lines = False)
			df_csv = df_csv.loc[:, ['INDEX'] + self.cols_name]
			# df_csv = self.checkError(df_csv)
			
			# if self.df.empty:
			# 	print(project_name)
			# 	index_num = self.df.loc[self.df.index == (project_name, test_ID), :].value_counts()
			# else:
			# 	index_num = 0						
			# df_csv.loc[:,'INDEX'] = index_num + df_csv['INDEX']
			# index_num += len(df_csv['INDEX'])

			# df_csv.insert(0, 'TIME', time)
			df_csv.insert(0, 'TEST', test_ID)
			df_csv.insert(0, 'PROJECT', project_name)
			df_csv = df_csv.set_index(['PROJECT', 'TEST', 'INDEX'])

			df_csv.columns = self.getNewColnames('PRE' if time == '0H' else 'POST')

			# self.df = pd.concat([self.df, df_csv], sort = False, axis = 1 )
			df_index = project_name + test_ID
			if df_index in df_dict.keys():
				df_dict[df_index] = pd.concat([df_dict[df_index], df_csv], axis = 1)
			else:
				df_dict[df_index] = df_csv

		for key,df in df_dict.items():
			self.df = pd.concat([self.df, df])
			# print(self.df)
			# print('------------------------------------------------------------------------------')

	def getNewColnames(self, second_col):

		new_col_list = []
		for colname in self.cols_name:
			new_col_list.append((colname, second_col))

		return pd.MultiIndex.from_tuples(new_col_list, names=['first', 'second'])

	def getNewColname(self, colname, second_col):
		return pd.MultiIndex.from_tuples((colname, second_col), names=['first', 'second'])

	def calRateByMode(self, num1, num2, mode):

		if mode not in ('RAW_PERCENT', 'STYLE_PERCENT', 'RAW_MULTI', 'STYLE_MULTI'):
			raise KeyError("The parameter Mod is NOT invalid")
		
		if '@F' in [num1, num2] or num1 == 0:
			print("Find a fail parameter")
			return '-'

		try:
			num1 = float(num1) if '@' not in num1 else float(num1[1:])
			num2 = float(num2) if '@' not in num2 else float(num2[1:])
		except TypeError:
			print("TYPE ERROR")
			return '-'

		if 'MULTI' in mode:
			return num2 / num1 if 'RAW' in mode else (f'{num2 / num1 :.2f}' + 'x')
		elif 'PERCENT' in mode:
			r = ((num2 - num1) / num1 )*100
			return r if 'RAW' in mode else ('+' if r > 0 else '') + f'{r :.2f}' + '%'
		else:
			print('MOD ERROR : ', mode)
			return '-'			

class CPDataObject():

	def __init__(self, file_list):

		self.wb = xlwt.Workbook()
		self.file_path = r"C:/Users/Wufanzhengshu/Desktop/JMP.xlsx"
		self.df = pd.DataFrame()
		self.cols_name = ['X', 'Y',			
				'BVDSS@VDSmax=1800V&IDS=200uA', 
				'VGS(TH)@IDS=28mA',
				'IDSS@VDS=1200V', 
				'IGSS@VGS=18V', 
				'RDS(ON)@IDS=100A&VG=18V']
		self.loadData(file_list)
		self.df.to_excel(self.file_path, sheet_name='JMP')

	def exportJMP(self):

		ws = self.wb.add_sheet('JMP', cell_overwrite_ok = True)
		file_list = []
		path = self.path

		for file in os.listdir(path):
			if 'BIN' in file or file[-3:] not in ['csv', 'CSV']:
				continue
			file_list.append(file)

		header = 11
		# cols_name = getColName(path+'/'+file_list[0], header)

		# ws.write(0, 0, 'WAFER')
		# ws.write(0, 1, 'X')
		# ws.write(0, 2, 'Y')
		for i in range(len(cols_name)):
			ws.write(0, i, cols_name[i])
		# for col in range(4, len(cols_name)):
		# 	ws.write(0, col-1, cols_name[col])

		row = 1
		for file in file_list:
			print(file)
			wafer = int(re.findall(r',(\d*)[,.]', file)[0])
			data_iter = csv.reader(open(path+'/'+file, 'r', encoding = 'gbk'))
			flag = 2

			while True:
				data = next(data_iter)
				if len(data) > 0 and '-' in data[0]:
					flag -= 1

				if flag == 0:
					break

			for data in data_iter:
				ws.write(row, 0, wafer)
				ws.write(row, 1, int(data[0]))
				ws.write(row, 2, int(data[1]))

				for col in range(len(col_num)):
					# print(data)
					if '@F' in data[col_num[col]]:
							ws.write(row, col + 3, '')
					elif '@' in data[col_num[col]]:
						try:
							ws.write(row, col + 3, float(data[col_num[col]][1:]))
						except:
							ws.write(row, col + 3, data[col_num[col]][1:])
					else:
						print('ERROR')

						continue
					# try:
					# 	ws.write(row, col + 3, float(data[col_num[col]]))
					# except ValueError:
						
				row += 1

		wb.save("C:/Users/Wufanzhengshu/Desktop/JMP2.xls")

	def getHeader(self, file):
		data_iter = csv.reader(open(file, 'r', encoding = 'gbk'))
		data = next(data_iter)
		header_len = 0
		flag = 2

		while flag > 0:
			if len(data) > 0 and '-' in data[0]:
				flag -= 1
			data = next(data_iter)
			header_len += 1

		return header_len

	def getColName(self, file, header):

		data_iter = csv.reader(open(file, 'r', encoding = 'gbk'))
		data = next(data_iter)
		index_dict = {	'No.' : [], 
						'Conditon1' : [], 
						'Conditon2' : [], 
						'Conditon3' : []}

		for i in range(0, header):
			if len(data) > 0 and data[0] in index_dict.keys():
				index_dict[data[0]] = data
			data = next(data_iter)

		index_dict['No.'][0] = 'X'
		index_dict['No.'][1] = 'Y'

		for i in range(4, len(index_dict['No.'])):

			if len(index_dict['Conditon1'][i]) > 0 and index_dict['Conditon1'][i] != '-':
				index_dict['No.'][i] = index_dict['No.'][i] + '@' + index_dict['Conditon1'][i]
			else:
				continue

			if index_dict['Conditon2'][i] != '-':
				index_dict['No.'][i] = index_dict['No.'][i] + '&' + index_dict['Conditon2'][i]

			if len(index_dict['Conditon3']) > 0 and index_dict['Conditon3'][i] != '-':
				index_dict['No.'][i] = index_dict['No.'][i] + '&' + index_dict['Conditon3'][i]	

		return index_dict['No.']

	def loadData(self, file_list):
		for fileIO in file_list:
			# print(file)
			file = fileIO.name
			wafer_ID = str(re.findall(r',(\w*)[,.]', file)[0])

			header_len = self.getHeader(file) 
			cols_name = self.getColName(file, header_len)

			for i in range(4, len(cols_name)):
				if cols_name[i] in cols_name[:i]:
					cols_name[i] += str(i)

			df_csv = pd.read_csv(file, header = header_len - 1, names = cols_name, skip_blank_lines = False)
			df_csv = df_csv.loc[:, self.cols_name]
			# df_csv = self.checkError(df_csv)

			df_csv.insert(0, 'WAFER', wafer_ID)

			self.df = pd.concat([self.df, df_csv], sort = False)


	def getRate(self):
		df = self.df.loc[self.df['WAFER'] == '09', ['X', 'Y']]
		for x in df['X']:
			for y in df.loc[df['X'] == x, 'Y']:
				pass

class WLBIDataObject:

	def __init__(self):
		pass

	def getColName(self, file):
		data_iter = csv.reader(open(file, 'r', encoding = 'gbk'))
		data = next(data_iter)
		current_row = 1

		while current_row <= 40:
			if data[0] == 'Test Name':
				print(data)

			if data[0] == 'Site#':
				print(data)
				break

			data = next(data_iter)

		return

def exportJMP(path, wb):

	ws = wb.add_sheet('JMP', cell_overwrite_ok = True)
	df = pd.DataFrame()
	file_list = os.listdir(path)

	for file in file_list:
		if file[-3:] not in ['csv', 'CSV']:
			file_list.remove(file)
			continue

	header_len = searchHeader(path+'/'+file_list[0])
	cols_name = getColName(path+'/'+file_list[0], header_len)

	ws.write(0, 0, 'INDEX')
	ws.write(0, 1, 'X')
	ws.write(0, 2, 'Y')
	for col in range(4, len(cols_name)):
		ws.write(0, col-1, cols_name[col])

	row = 1
	for file in file_list:
		df_csv = pd.DataFrame()
		index = int(re.findall(r',(\d*)[,.]', file)[0])
		data_iter = csv.reader(open(path+'/'+file, 'r', encoding = 'gbk'))

		for i in range(0, header_len):
			next(data_iter)

		for data in data_iter:
			ws.write(row, 0, index)
			ws.write(row, 1, int(data[0]))
			ws.write(row, 2, int(data[1]))
			for col in range(4, len(data)):
				try:
					ws.write(row, col-1, float(data[col]))
				except ValueError:
					ws.write(row, col-1, data[col])
			row += 1

	wb.save("C:/Users/Wufanzhengshu/Desktop/result.xls")

if __name__ == '__main__' :
	print("------------------------------------------------------")

	cols_name = [			
		'BVDSS@VDSmax=1800V&IDS=200uA', 
		'VGS(TH)@IDS=10mA',
		'IDSS@VDS=1200V', 
		'IGSS@VGS=18V', 
		'RDS(ON)@IDS=40A&VG=15V']
	# cols_name = ['BVR@VRmax=1800V&IR=200uA', 'IR@VR=1200V', 'VF@IF=25A']

	path = filedialog.askdirectory()
	data = DataObject(cols_name = cols_name)
	data.loadFromFiles(path)
	# data.addData(r'C:\Users\Wufanzhengshu\Desktop\3193\Data.xlsx')

	data.save(path, 'Data.xlsx')


	# wlbi = WLBIDataObject()
	# wlbi.getColName(r'D:\测试数据\Wafer Burnin\WLBI DATA\1557-02_WLBI01_20230802_164318\\CSV\1557-02__WLBI01_173151_20230802.csv')