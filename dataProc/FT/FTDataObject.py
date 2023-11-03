#-- coding: utf-8 --
import pandas as pd
import openpyxl, re, os, configparser, json
import plotly.graph_objects as go
from datetime import datetime
from numpy import nan

import Configuration
LOT = Configuration.LOT_FLAG
DEVICE = Configuration.DEVICE_FLAG
COL_FLAG = Configuration.COL_FLAG

date = datetime.today()

class FTDataObject():

	def __init__(self, device):

		self.HEADER = 64
		# self.standard_list = ['INDEX', 'BIN', 'IDSS', 'VTH', 'BV', 'RDON', 'IGSS', 'VF', 'VZ', 'IR']
		self.standard_list = Configuration.TEST_KEY
		self.df = pd.DataFrame()
		self.lot_list = []
		self.device = device
		self.filename_mod =  Configuration.FILENAME_MOD
		# cfg = configparser.ConfigParser()
		# cfg.read('./Configuration.ini')
		# self.filename_mod = cfg.get('setting', 'filename_mod')

	def getStatistic(self):
		df = self.df
		lots = df[LOT].unique()
		statistic_dict = {}
		for lot in lots:
			
			one_df = df.loc[df[LOT] == lot,:]
			bin_series = one_df['HARD_BIN'].value_counts()
			bin_sum = bin_series.sum()
			statistic_dict[lot] = { 
									'COUNT': bin_sum,
									'GOOD': bin_series[2]/bin_sum
									}
		return statistic_dict

	def _dfReader(self, file_path, header=0, nrows = None):
		if file_path.endswith(('.csv', '.CSV')):
			return pd.read_csv(file_path, header = header, nrows = nrows )#, delimiter='\t', encoding='gbk')
		elif file_path.endswith(('.xls')):
			return pd.read_excel(file_path, header = header, nrows = nrows)
		else:
			return DataFrame()

	def dataClean(self):
		df = self.df
		for colname in df.columns:
			if 'IGSS' not in colname:
				continue

			min_mask = df[colname] < df[colname].median() * 1000
			max_mask = df[colname] > df[colname].median() * 1000
			df.loc[min_mask|max_mask, colname] = None

		self.df = df

	def scanFilename(self, file_name):
		#根据文件名获取文件信息（LOT号、器件号）
		file_name = os.path.basename(file_path)
		info_dict = self.getInfoFromFilename(file_name)
		# device = re.findall(r'(S\w+[M|D]\d+[A|H]?D?)',file_name)[0]
		# lot = re.findall(r'([B|BP]\d*)',file_name)[0]
		device = info_dict[DEVICE]
		lot = info_dict[LOT]

	def getData(self, file_info):

		if type(file_info) == dict:
			file_path = file_info['PATH']
			lot = file_info[LOT]
			device = file_info[DEVICE]
		elif type(file_info) == str:
			file_path = file_info
			file_name = os.path.basename(file_info)
			device = re.findall(r'(SG?\d+[M|D]\d+[A|H]?D?\d?)',file_name)[0]
			lot = re.findall(r'([B|BP]\d*[.]?\d*)',file_name)[0]
		else:
			raise TypeError("Function getData need receive a dict or str, please check the file type.")

		self.file_path = file_path
		header = self.searchHeader(file_path)
		if not header:
			print(f'[Value Error]: Do NOT search the header of the file({file_path})')
			return None

		# df = pd.read_excel(file_path, header = header)
		df = self._dfReader(file_path, header = header)

		#修改第一列列名为INDEX
		df = df.rename(columns={COL_FLAG:'INDEX'})
		df = df.drop(index=df.index[0:7])

		#删除冗余行列
		del_list = []
		for col in df.columns:
			if all(standard not in col for standard in self.standard_list):
				del_list.append(col)
		df = df.drop(del_list, axis = 1)
		
		# df.loc[:, ~df.columns.isin(['INDEX', 'HARD_BIN'])].apply(pd.to_numeric, errors='ignore')
		df = df.apply(pd.to_numeric, errors='ignore')
		df.insert(0, LOT, lot)
		df.insert(0, DEVICE, device)

		pd.set_option('display.max_columns', None)

		return df

	def addDF(self, one_df):
		if type(one_df) != pd.DataFrame:
			return

		lot = one_df[LOT].unique()[0]

		

		if not self.df.empty and lot in self.df[LOT].unique():
			count = len(self.df.loc[self.df[LOT] == lot,	 'INDEX'])
			one_df['INDEX'] = one_df['INDEX'] + count		
			print(f"[Insert Data]: LOT({lot}) already exists, inserting data starting from index {count+1}")

		if lot not in self.lot_list:
			# print('Add a LOT', lot)
			self.lot_list.append(lot)

		self.df = pd.concat([self.df, one_df])

	def loadDatabyDict(self, data_info):
		if type(data_info) != dict:
			raise TypeError("Function loadDatabyDict need to accept a dict parameter, please check the file type.")

		one_df = self.getData(data_info)
		print(f"[Load Data]: Loading new data from file({data_info['PATH']})")
		self.addDF(one_df)

		return self.df

	def getUnits(self):

		file = self.file_path
		with open(file) as f:
			
			for i in range(100):
				line = f.readline()
				if line == '\n':
					continue
				row = line.split(',')
				if row[0] == COL_FLAG:
					title_list = row
				if row[0] == 'Units':
					# print("Find header", header)
					unit_list = row
					break
		unit_dict = {}
		for index in range(1, len(title_list)):
			if len(unit_list[index]) == 0:
				continue

			if title_list[index] in self.df.columns:
				unit_dict[title_list[index]] = unit_list[index]

		self.units = unit_dict
		return unit_dict

	def searchHeader(self, file):

		header = 0
		with open(file) as f:
			
			for i in range(100):
				line = f.readline()
				if line == '\n':
					continue
				row = line.split(',')
				if row[0] == COL_FLAG:
					# print("Find header", header)
					return header
				header += 1

		return None

class FTReport():

	def __init__(self, wb = None):
		if type(wb) is openpyxl.workbook.workbook.Workbook:
			self.wb = wb
		else:	
			self.wb = openpyxl.Workbook()
			self.wb.create_sheet('Summary')
		# self.wb = wb
		self.createDir()

	def getParaList(self, ftd):
		key_list = ['IDSS', 'VTH', 'BV', 'RDON', 'IGSS', 'VF', 'VZ', 'IR']
		colunm_list = list(ftd.df.columns)
		para_list = []

		for col in colunm_list:
			if any(key in col for key in key_list):
				para_list.append(col)

		return para_list

	def getImgPara(self, device):
		conf = configparser.ConfigParser()
		conf.read('./Figure parameter.ini', encoding="utf-8-sig")

		if device in conf.sections():
			return [conf.get(device, 'SCATTER_X'), conf.get(device, 'SCATTER_Y')]
		else:
			return [conf.get('default', 'SCATTER_X'), conf.get('default', 'SCATTER_Y')]

	def addSummary(self, ftd):
		 
		# list_head = ['DEVICE', 'LOT', 'COUNT', 'CP Yield','GS Short', 'DS Leakage', 'Vth bad', 'Bady Diode Bad', 'Rdson Bad']
		# BIN = {
		# 	'CP Yield':2,
		# 	'GS Short':3, 
		# 	'DS Leakage':4, 
		# 	'Vth bad':5, 
		# 	'Bady Diode Bad':6, 
		# 	'Rdson Bad':7
		# }
		list_head = Configuration.SUMMARY_LIST
		bin_dict = Configuration.BIN_DICT 

		df = ftd.df
		lot_list = ftd.lot_list
		device = ftd.device
		# print(device, lot_list)
		print(f'[ADD SUMMARY] Adding summary data ({device}) to the sheet , including {lot_list}')
		# Create a new sheet named Summary if there is no this sheet.
		if 'Summary' in self.wb.sheetnames:
			ws = self.wb['Summary']
		else:
			ws = self.wb.create_sheet('Summary')

		row = ws.max_row + 2 if ws.max_row != 1 else 1

		# lot_list = df.loc[df['DEVICE'] == device, LOT].unique()
		
		para_list = self.getParaList(ftd)
		title_list = list_head + para_list
		for col, title in enumerate(title_list):
			ws.cell(row=row, column=col+1, value=title)
		
		for lot in lot_list:
			row += 1
			one_df = df.loc[df[LOT] == lot,:]
			bin_series = one_df['HARD_BIN'].value_counts()
			bin_sum = bin_series.sum()
			ws.cell(row=row, column=1, value=device)
			ws.cell(row=row, column=2, value=lot)
			ws.cell(row=row, column=3, value=bin_sum)
	
			for key, code in bin_dict.items():
				if code in bin_series:
					value = '{}%'.format((bin_series[code]/bin_sum * 100).round(2))
				else:
					value = '-'
				ws.cell(row=row, column=list_head.index(key)+1, value=value)

			for col, parameter in enumerate(para_list):
				median = one_df[parameter].median()
				ws.cell(row=row, column=col+len(list_head)+1, value=median)

	def _drawBoxplot(self, x_data, y_data, tle):
		colors = ['rgba(93, 164, 214, 0.5)', 'rgba(255, 144, 14, 0.5)', 'rgba(44, 160, 101, 0.5)',
          'rgba(255, 65, 54, 0.5)', 'rgba(207, 114, 255, 0.5)', 'rgba(127, 96, 0, 0.5)', 'rgba(93, 164, 214, 0.5)', 'rgba(255, 144, 14, 0.5)', 'rgba(44, 160, 101, 0.5)',
          'rgba(255, 65, 54, 0.5)', 'rgba(207, 114, 255, 0.5)', 'rgba(127, 96, 0, 0.5)']
		file_path = os.path.join(self.Img_path, tle + '.png')
		fig = go.Figure()

		for xd, yd, cls in zip(x_data, y_data, colors):
			fig.add_trace(go.Box(
	            y=yd,
	            name=xd,
	            boxpoints='all',
	            jitter=0.5,
	            whiskerwidth=0.2,
	            # fillcolor=cls,
	            marker_size=2,
	            line_width=1)
	        )

		fig.update_layout(
		    title=tle,

		    yaxis=dict(
		        autorange=True,
		        showgrid=True,
		        zeroline=True,

		        # dtick=2,

		        gridcolor='rgb(255, 255, 255)',
		        gridwidth=1,

		        zerolinecolor='rgb(255, 255, 255)',
		        zerolinewidth=2,
		    ),

		    margin=dict(
		        l=40,
		        r=30,
		        b=40,
		        t=50,
		    ),

		    paper_bgcolor='rgb(243, 243, 243)',
		    plot_bgcolor='rgb(243, 243, 243)',
		    showlegend=False
		)

		fig.write_image(file_path, engine="kaleido", scale = 1.3)
		return file_path
		# fig.show()

	def _drawScatter(self, x_data, y_data, tle):
		fig = go.Figure()
		file_path = os.path.join(self.Img_path, tle + '.png')

		fig.add_trace(go.Scatter(
		    x = x_data,
		    y = y_data,
		    mode='markers'
		))

		fig.update_layout(
			    title= tle,

			    yaxis=dict(
			        autorange=True,
			        showgrid=True,
			        zeroline=True,
			        
			        # dtick=dtick,

			        gridcolor='rgb(255, 255, 255)',
			        gridwidth=1,

			        zerolinecolor='rgb(255, 255, 255)',
			        zerolinewidth=2,
			    ),

			    margin=dict(
			        l=40,
			        r=30,
			        b=40,
			        t=50,
			    ),

			    paper_bgcolor='rgb(243, 243, 243)',
			    plot_bgcolor='rgb(243, 243, 243)',
			    showlegend=False
			)

		fig.write_image(file_path, engine="kaleido", scale = 1.3)
		return file_path

	def createImgSheet(self, ftd):
		ws = self.wb.create_sheet(ftd.device)

		boxplot_list = self._addBoxplot(ftd)
		fig_step = 0	
		for img in boxplot_list:
			img_stream = openpyxl.drawing.image.Image(img)
			ws.add_image(img_stream, 'A' + str(1 + fig_step))
			fig_step += 38

		try:
			scatter_list = self._addScatter(ftd)
		except:
			print(f'[Draw ERROR]: There is an error in the process of drawing scatter (device:{ftd.device})')
			scatter_list = []

		fig_step = 0	
		for img in scatter_list:
			img_stream = openpyxl.drawing.image.Image(img)
			ws.add_image(img_stream, 'Q' + str(1 + fig_step))
			fig_step += 38

	def _addBoxplot(self, ftd):
		df = ftd.df
		lot_list = ftd.lot_list
		device = ftd.device
		para_list = self.getParaList(ftd)

		boxplot_data = {}
		fig_list = []
		for lot in lot_list:
			for para in para_list:
				if para not in boxplot_data.keys():
					boxplot_data[para] = []

				if 'IGSS' not in para:
					data_np = df.loc[(df[LOT] == lot), para]
				else:
					median = df[para].median()
					mask = df[para] < median * 1000 if median > 0 else df[para] > median * 1000
					# max_mask = df[para] < median * 1000
					# min_mask = df["weight"] < (df["weight"].mean() - 3 *  df["weight"].std())
					# max_mask = df["weight"] > (df["weight"].mean() + 3 * df["weight"].std())

					data_np = df.loc[(df[LOT] == lot) & mask, para]

				boxplot_data[para].append(list(data_np))

		for key, data in boxplot_data.items():

			if key in ftd.units.keys():
				unit = f' ({ftd.units[key]})'
			else:
				unit = ''

			try:
				img_path = self._drawBoxplot(lot_list, data, device + '-' + key + unit)
				print(f'[Draw Figure]: Drawing the Boxplot figure of {key}')
				fig_list.append(img_path)
			except:
				print(f'[ERROR]: Fail to drawing the Boxplot figure of {key}')

		
		return fig_list
			
	def _addScatter(self, ftd):
		df = ftd.df
		lot_list = ftd.lot_list
		device = ftd.device
		para_list = self.getParaList(ftd)
		scatter_X, scatter_Y = self.getImgPara(device)
	
		scatter_data = {}
		fig_list = []
		for lot in lot_list:
			scatter_data[lot] = {
				'X_DATA': df.loc[(df[LOT] == lot), scatter_X], 
				'Y_DATA': df.loc[(df[LOT] == lot), scatter_Y], 
			}

			img_path = self._drawScatter(
				df.loc[(df[LOT] == lot), scatter_X], 
				df.loc[(df[LOT] == lot), scatter_Y], 
				lot)
			print(f'[Draw Figure]: Drawing the VTH-RDSON scatter figure of lot({lot})')
			fig_list.append(img_path)

		return fig_list

	def createDir(self):

		dir_path = os.path.join(Configuration.SAVE_DIR, date.strftime('%Y-%m-%d %H_%M_%S'))
		img_path = os.path.join(dir_path, 'IMG')

		if not os.path.exists(dir_path):
			os.mkdir(dir_path)
			os.mkdir(img_path)
		elif not os.path.exists(img_path):
			os.mkdir(img_path)

		self.save_path = dir_path
		self.Img_path = img_path

	def save(self):
		if 'Sheet' in self.wb.sheetnames:
			del self.wb['Sheet']

		timestamp = date.strftime('%Y%m%d_%H%M%S')
		save_path = os.path.join(self.save_path, f'FTReport_{timestamp}.xlsx')
		self.wb.save(save_path)
		return save_path

def main():

	FTD_dict = {}
	summary_dict = {}
	data_folder = r'Z:\2023\生产数据'
	fd = FileData()
	file_list = fd.getFiles(data_folder)

	if len(file_list) == 0:
		print("No File Be Found")
		return None

	for info in file_list:

		device = info[DEVICE]

		file_path = info['PATH']
		if device not in FTD_dict.keys():
			ftd = FTDataObject(device)
			FTD_dict[device] = ftd
		else:
			ftd = FTD_dict[device]
		ftd.loadDatabyDict(info)
		# ftd.dataClean()

	report = FTReport()

	for device, ftd in FTD_dict.items():
		u = ftd.getUnits()
		# print(json.dumps(u, indent=4, ensure_ascii=False, sort_keys=False))
		summary_dict[device] = ftd.getStatistic()
		# print(json.dumps(a, indent=4, ensure_ascii=False, sort_keys=False))
		report.addSummary(ftd)
		
		report.createImgSheet(ftd)


	report_path = report.save()
	# # report_path = r'C:\Users\Wufanzhengshu\Desktop\FT测试报告\2023-10-09 16_11_45\FTReport_20231009_161258.xlsx'
	mail_content = emailEdit(summary_dict)
	print(mail_content)
	# aliMailSender(mail_content, report_path)
	# fd.updateConfTime()

if __name__ == '__main__':

	main()

		




		